
---

### 2. report.py  
これは「実際に動くコード」を書く場所です。  
古いファイル削除の処理などはこちらにまとめて入れます。

```python
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference
import os
import glob

# ダミーデータ作成
data = {
    "Date": pd.date_range(start="2025-01-01", periods=7, freq="D"),
    "Sales": [1200, 1500, 900, 1800, 2000, 1700, 2200],
    "Region": ["Tokyo", "Osaka", "Nagoya", "Tokyo", "Osaka", "Nagoya", "Tokyo"]
}
df = pd.DataFrame(data)
summary = df.groupby("Region")["Sales"].sum().reset_index()

# Excel出力
filename = f"sales_report_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
with pd.ExcelWriter(filename, engine="openpyxl") as writer:
    df.to_excel(writer, sheet_name="Raw Data", index=False)
    summary.to_excel(writer, sheet_name="Summary", index=False)

# グラフ追加
wb = load_workbook(filename)
ws = wb["Summary"]
chart = BarChart()
chart.title = "Sales by Region"
data_ref = Reference(ws, min_col=2, min_row=1, max_row=ws.max_row)
cats_ref = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
chart.add_data(data_ref, titles_from_data=True)
chart.set_categories(cats_ref)
ws.add_chart(chart, "D2")
wb.save(filename)

print(f"[OK] Excel report (with chart) saved: {filename}")

# 古いレポートを削除（最新1件だけ残す）
files = sorted(glob.glob("sales_report_*.xlsx"))
if len(files) > 1:
    for f in files[:-1]:
        os.remove(f)
        print(f"[INFO] Deleted old report: {f}")

